import tika
import PyPDF2
import os
import openpyxl
import sys
from tika import parser

while True:
    try:
        print('This will scrape circuit numbers from a PDF. It requires a list of all panel names in the job.\nUse only power drawings you want circuits from.\n')
        plans = input('Enter the name of the PDF plans to be scraped: ')
        plans = plans + '.pdf'
        break
    except:
        print('That .pdf file was not found. Try again.\n')

while True:
    try:
        minLabelNumber = input('Enter the minimum number of labels to create for each circuit: ')
        minLabelNumber = int(minLabelNumber)
        break
    except:
        print('Please enter a whole number only.\n')

while True:
    try:
        panelNameFile = input('Enter the name of the .txt file with all panel names, exported from Revit: ')
        panelNameFile = panelNameFile + ('.txt')
        break
    except:
        print('That .txt file was not found. Try again.\n')


raw = parser.from_file(plans)
content = raw['content']
split = content.split('\n')
circuits = []
panelNames = []

with open(panelNameFile) as file:
    lines = file.readlines()[2:]
    for line in lines:
        line = line.strip()
        line = line.strip('"')
        if line is not '':
            panelNames.append(line)

for item in split:
    if item.startswith(tuple(panelNames)) and '-' in item:
        itemSplit = item.split()
        for i in itemSplit:
            if i.startswith(tuple(panelNames)) and '-' in i:
                circuits.append(i)

file = open(plans, 'rb')
pdfRotator = PyPDF2.PdfFileReader(file)
pdfWriter = PyPDF2.PdfFileWriter()
for pageNumber in range(pdfRotator.numPages):
    page = pdfRotator.getPage(pageNumber)
    page.rotateClockwise(90)
    pdfWriter.addPage(page)

rotatedPdf = open('rotated.pdf', 'wb')
pdfWriter.write(rotatedPdf)
rotatedPdf.close()

raw = parser.from_file('rotated.pdf')
content = raw['content']
split = content.split('\n')

for item in split:
    if item.startswith(tuple(panelNames)) and '-' in item:
        itemSplit = item.split()
        for i in itemSplit:
            if i.startswith(tuple(panelNames)) and '-' in i and i not in circuits:
                circuits.append(i)

circuits.sort()
os.remove('rotated.pdf')

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'Sheet1'
nextRow = 1

for circuit in circuits:
    sheet.cell(row = nextRow, column = 1, value = circuit)
    sheet.cell(row = nextRow, column = 2, value = minLabelNumber)
    nextRow += 1

wb.save('circuits.xlsx')